"""
Professional Excel Report Generator using OpenPyXL
Generates styled, executive-ready Excel reports for bulk domain threat queries.
"""

from typing import Dict
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(report_data: Dict, file_path: str):
    """
    Generates a multi-sheet styled Excel report.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # STYLES DEFINITIONS
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    # Header Fonts & Fills
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="E2E8F0")
    section_font = Font(name=font_family, size=12, bold=True, color="0F172A")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    kpi_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    
    # Card Fills
    card_blocked_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    card_clean_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    card_info_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Badge Fills & Fonts
    blocked_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    blocked_font = Font(name=font_family, size=10, bold=True, color="991B1B")
    
    clean_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    clean_font = Font(name=font_family, size=10, bold=True, color="065F46")
    
    error_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    error_font = Font(name=font_family, size=10, bold=True, color="92400E")
    
    data_font = Font(name=font_family, size=10, color="334155")
    mono_font = Font(name="Consolas", size=9, color="0F172A")
    
    # Zebra Fills
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Borders
    thin_side = Side(style="thin", color="CBD5E1")
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom = Border(bottom=Side(style="medium", color="0F172A"))
    
    summary = report_data.get("summary", {})
    results = report_data.get("results", [])
    
    # ----------------------------------------------------
    # SHEET 1: ÖZET RAPOR (SUMMARY DASHBOARD)
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Özet Rapor"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Banner Header
    ws_summary.merge_cells("B2:G3")
    banner_cell = ws_summary["B2"]
    banner_cell.value = "T.C. SİBER GÜVENLİK BAŞKANLIĞI"
    banner_cell.font = title_font
    banner_cell.fill = header_fill
    banner_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for r in range(2, 4):
        for c in range(2, 8):
            cell = ws_summary.cell(row=r, column=c)
            cell.fill = header_fill
    
    ws_summary["B4"] = "Toplu Alan Adı Tehdit İstihbaratı ve Engelleme Sorgu Raporu"
    ws_summary["B4"].font = Font(name=font_family, size=11, bold=True, color="475569")
    
    # KPI Cards Table
    ws_summary["B6"] = "GENEL TARAMA ÖZETİ"
    ws_summary["B6"].font = section_font
    
    kpis = [
        ("Tarama Tarihi", summary.get("scan_timestamp", "-"), card_info_fill),
        ("Toplam Sorgulanan Domain", summary.get("total_scanned", 0), card_info_fill),
        ("Engellenmiş / Zararlı Domain", summary.get("blocked_count", 0), card_blocked_fill),
        ("Temiz / Güvenli Domain", summary.get("clean_count", 0), card_clean_fill),
        ("Kritik Seviye (Seviye 1-3)", summary.get("high_criticality_count", 0), card_blocked_fill),
        ("Tarama Süresi (Saniye)", summary.get("duration_seconds", 0), card_info_fill),
        ("Tarama Hızı (Domain/Sn)", summary.get("domains_per_second", 0), card_info_fill),
    ]
    
    row_idx = 8
    ws_summary.cell(row=row_idx, column=2, value="Metrik").font = header_font
    ws_summary.cell(row=row_idx, column=2).fill = kpi_header_fill
    ws_summary.cell(row=row_idx, column=3, value="Değer").font = header_font
    ws_summary.cell(row=row_idx, column=3).fill = kpi_header_fill
    
    for metric, val, fill in kpis:
        row_idx += 1
        cell_m = ws_summary.cell(row=row_idx, column=2, value=metric)
        cell_v = ws_summary.cell(row=row_idx, column=3, value=val)
        
        cell_m.font = Font(name=font_family, size=10, bold=True, color="334155")
        cell_v.font = Font(name=font_family, size=11, bold=True, color="0F172A")
        cell_m.fill = fill
        cell_v.fill = fill
        cell_m.border = border_thin
        cell_v.border = border_thin
        cell_v.alignment = Alignment(horizontal="right")

    # Pie Chart
    pie = PieChart()
    pie.title = "Domain Dağılım Grafiği"
    data_ref = Reference(ws_summary, min_col=3, min_row=10, max_row=11)
    labels_ref = Reference(ws_summary, min_col=2, min_row=10, max_row=11)
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(labels_ref)
    pie.width = 14
    pie.height = 7.5
    ws_summary.add_chart(pie, "E8")

    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 25
    ws_summary.column_dimensions["D"].width = 5

    # ----------------------------------------------------
    # SHEET 2: DETAYLI LİSTE (DATA TABLE)
    # ----------------------------------------------------
    ws_data = wb.create_sheet(title="Detaylı Liste")
    ws_data.views.sheetView[0].showGridLines = True
    
    headers = [
        "Sıra",
        "Domain Adı",
        "Durum",
        "Eşleşme Türü",
        "En Yüksek Kritiklik",
        "Kayıt Sayısı",
        "Tehdit Kategorisi",
        "Tehdit Kaynağı",
        "Bağlantı Tipi",
        "Kayıt / Tespit Tarihi",
        "Açıklama / Notlar"
    ]
    
    # Header Row
    ws_data.row_dimensions[1].height = 28
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data Rows
    for idx, item in enumerate(results, start=1):
        row_num = idx + 1
        ws_data.row_dimensions[row_num].height = 22
        
        domain_name = item.get("domain", "")
        status = item.get("status", "BİLİNMİYOR")
        match_type = item.get("match_type", "-")
        max_crit = item.get("max_criticality", "-")
        match_count = item.get("match_count", 0)
        
        threats = item.get("threats", [])
        if threats:
            t0 = threats[0]
            cat = t0.get("desc_text", "-")
            src = t0.get("source_text", "-")
            conn = t0.get("connection_type_text", "-")
            dt = t0.get("date", "-")
            desc_detail = f"ID: {t0.get('id')} | Toplam {len(threats)} Tehdit Kaydı"
        else:
            cat = "-"
            src = "-"
            conn = "-"
            dt = "-"
            desc_detail = item.get("error") or "Zararlı kaydı bulunamadı (Güvenli)"

        row_data = [
            idx,
            domain_name,
            status,
            match_type,
            max_crit if max_crit is not None else "-",
            match_count,
            cat,
            src,
            conn,
            dt,
            desc_detail
        ]

        is_even = (idx % 2 == 0)

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")
            
            # Default Zebra Background
            if is_even:
                cell.fill = zebra_fill
            
            # Custom formatting per column
            if col_idx == 1:  # Index
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:  # Domain
                cell.font = mono_font
            elif col_idx == 3:  # Status Badge
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if status == "ENGELLENMİŞ":
                    cell.fill = blocked_fill
                    cell.font = blocked_font
                elif status == "TEMİZ":
                    cell.fill = clean_fill
                    cell.font = clean_font
                else:
                    cell.fill = error_fill
                    cell.font = error_font
            elif col_idx in (5, 6):  # Criticality, Match count
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 10:  # Date
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col in ws_data.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        adjusted_width = max(max_len + 4, 12)
        if adjusted_width > 50:
            adjusted_width = 50
        ws_data.column_dimensions[col_letter].width = adjusted_width

    # Save workbook
    wb.save(file_path)
